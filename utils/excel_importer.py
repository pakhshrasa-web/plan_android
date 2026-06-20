"""
وارد کردن اطلاعات از فایل Excel - نسخه بهینه برای اندروید
"""

import os
import openpyxl
from utils.file_manager import add_route, add_customer, get_routes, get_customers

def import_routes_from_excel(filepath):
    """
    وارد کردن مسیرها از فایل Excel
    فرمت فایل: ستون اول = name
    """
    if not os.path.exists(filepath):
        return False, "فایل وجود ندارد"
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        imported_count = 0
        duplicate_count = 0
        existing_routes = [r.get('name', '') for r in get_routes()]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
                
            name = str(row[0]).strip()
            if name:
                if name not in existing_routes:
                    add_route({'name': name})
                    imported_count += 1
                    existing_routes.append(name)
                else:
                    duplicate_count += 1
        
        wb.close()
        return True, f"{imported_count} مسیر جدید وارد شد. {duplicate_count} مسیر تکراری نادیده گرفته شد."
    
    except Exception as e:
        return False, f"خطا در خواندن فایل: {str(e)}"

def import_customers_from_excel(filepath):
    """
    وارد کردن مشتریان از فایل Excel
    فرمت فایل: name, store_name, route_name, mobile, address
    """
    if not os.path.exists(filepath):
        return False, "فایل وجود ندارد"
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        imported_count = 0
        duplicate_count = 0
        error_count = 0
        routes = get_routes()
        route_names = [r.get('name', '') for r in routes]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                error_count += 1
                continue
            
            # خواندن مقادیر با مدیریت None
            name = str(row[0]).strip() if row[0] else ''
            store_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            route_name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            mobile = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            address = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            
            if not name:
                error_count += 1
                continue
            
            # بررسی تکراری نبودن
            existing = get_customers()
            is_duplicate = any(c.get('name') == name for c in existing)
            
            if not is_duplicate:
                # بررسی وجود مسیر
                if route_name and route_name not in route_names:
                    route_names.append(route_name)
                    add_route({'name': route_name})
                
                customer = {
                    'name': name,
                    'store_name': store_name,
                    'route_name': route_name,
                    'mobile': mobile,
                    'address': address
                }
                add_customer(customer)
                imported_count += 1
            else:
                duplicate_count += 1
        
        wb.close()
        
        msg = f"{imported_count} مشتری جدید وارد شد."
        if duplicate_count > 0:
            msg += f" {duplicate_count} مشتری تکراری نادیده گرفته شد."
        if error_count > 0:
            msg += f" {error_count} سطر خطا داشت."
        
        return True, msg
    
    except Exception as e:
        return False, f"خطا در خواندن فایل: {str(e)}"

def get_excel_template_routes():
    """ایجاد فایل نمونه برای مسیرها"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Routes"
    ws['A1'] = "name"
    ws['A2'] = "نمونه: مسیر شمال"
    return wb

def get_excel_template_customers():
    """ایجاد فایل نمونه برای مشتریان"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws['A1'] = "name"
    ws['B1'] = "store_name"
    ws['C1'] = "route_name"
    ws['D1'] = "mobile"
    ws['E1'] = "address"
    
    ws['A2'] = "علی محمدی"
    ws['B2'] = "فروشگاه رضا"
    ws['C2'] = "مسیر شمال"
    ws['D2'] = "09121234567"
    ws['E2'] = "خیابان ولیعصر"
    return wb

def import_from_excel(filepath, data_type='customers'):
    """
    تابع عمومی برای وارد کردن از اکسل
    data_type: 'customers' یا 'routes'
    """
    if data_type == 'routes':
        return import_routes_from_excel(filepath)
    elif data_type == 'customers':
        return import_customers_from_excel(filepath)
    else:
        return False, "نوع داده نامعتبر است"