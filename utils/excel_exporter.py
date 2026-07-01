"""
ساخت خروجی Excel از ویزیت‌ها - نسخه بهینه برای اندروید
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from utils.file_manager import get_daily_logs, get_data_path


def safe_int(value, default=0):
    """تبدیل امن به عدد صحیح"""
    if value is None:
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def safe_str(value, default=''):
    """تبدیل امن به رشته"""
    if value is None:
        return default
    return str(value)


def export_to_excel():
    """
    خروجی گرفتن از تمام ویزیت‌ها به فایل Excel
    برگرداندن مسیر کامل فایل ایجاد شده
    """
    try:
        data_path = get_data_path()
        all_logs = get_daily_logs()
        
        if not all_logs:
            return None
        
        wb = Workbook()
        
        # ========== صفحه اول: گزارش ویزیت‌ها (همه ردیف‌ها) ==========
        ws1 = wb.active
        ws1.title = "گزارش ویزیت‌ها"
        
        # استایل‌ها
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # هدرها
        headers = ["ردیف", "تاریخ", "مسیر", "مشتری", "وضعیت ویزیت", 
                   "وضعیت فروش", "تعداد واحد", "مبلغ فروش", "نحوه تسویه", "ساعت"]
        
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # پر کردن داده‌ها - همه ردیف‌ها
        row = 2
        total_sales = 0
        total_invoices = 0
        total_visits = 0
        total_units = 0
        total_successful_visits = 0
        total_failed_visits = 0
        total_successful_sales = 0
        total_failed_sales = 0
        
        sorted_dates = sorted(all_logs.keys(), reverse=True)
        idx = 1
        
        for date in sorted_dates:
            log_list = all_logs[date]
            if not isinstance(log_list, list):
                continue
            
            for log in log_list:
                if not isinstance(log, dict):
                    continue
                
                visit_status = log.get('visit_status', '')
                sales_status = log.get('sales_status', '')
                sales_amount = safe_int(log.get('sales_amount', 0))
                units_sold = safe_int(log.get('units_sold', 0))
                
                ws1.cell(row=row, column=1, value=idx)
                ws1.cell(row=row, column=2, value=safe_str(date))
                ws1.cell(row=row, column=3, value=safe_str(log.get('route', '')))
                ws1.cell(row=row, column=4, value=safe_str(log.get('customer', '')))
                ws1.cell(row=row, column=5, value=safe_str(visit_status))
                ws1.cell(row=row, column=6, value=safe_str(sales_status if sales_status else '---'))
                ws1.cell(row=row, column=7, value=units_sold if sales_status == 'موفق' else 0)
                ws1.cell(row=row, column=8, value=sales_amount if sales_status == 'موفق' else 0)
                ws1.cell(row=row, column=9, value=safe_str(log.get('payment_method', '---')))
                ws1.cell(row=row, column=10, value=safe_str(log.get('time', '')))
                
                # آمار
                total_visits += 1
                if visit_status == 'موفق':
                    total_successful_visits += 1
                    if sales_status == 'موفق':
                        total_successful_sales += 1
                        total_invoices += 1
                        total_units += units_sold
                        total_sales += sales_amount
                    else:
                        total_failed_sales += 1
                else:
                    total_failed_visits += 1
                
                row += 1
                idx += 1
        
        # تنظیم عرض ستون‌ها
        col_widths = [8, 14, 18, 22, 14, 14, 14, 18, 14, 14]
        for col, width in enumerate(col_widths, 1):
            ws1.column_dimensions[get_column_letter(col)].width = width
        
        # ========== صفحه دوم: خلاصه آمار ==========
        ws2 = wb.create_sheet("خلاصه آمار")
        
        summary_data = [
            ["شاخص", "مقدار"],
            ["کل فروش (ریال)", f"{total_sales:,}"],
            ["تعداد کل فاکتورها", total_invoices],
            ["تعداد کل ویزیت‌ها", total_visits],
            ["تعداد ویزیت موفق", total_successful_visits],
            ["تعداد ویزیت ناموفق", total_failed_visits],
            ["تعداد فروش موفق", total_successful_sales],
            ["تعداد فروش ناموفق", total_failed_sales],
            ["تعداد کل واحد فروش", total_units],
            ["تعداد روزهای کاری", len(all_logs)],
            ["میانگین مبلغ هر فاکتور", f"{total_sales // total_invoices:,}" if total_invoices > 0 else "۰"],
            ["میانگین فروش هر ویزیت موفق", f"{total_sales // total_successful_visits:,}" if total_successful_visits > 0 else "۰"],
        ]
        
        for r, row_data in enumerate(summary_data, 1):
            for c, value in enumerate(row_data, 1):
                cell = ws2.cell(row=r, column=c, value=value)
                if r == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="28B463", end_color="28B463", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 25
        
        # ========== صفحه سوم: آمار روزانه ==========
        ws3 = wb.create_sheet("آمار روزانه")
        
        daily_headers = ["تاریخ", "کل ویزیت", "ویزیت موفق", "ویزیت ناموفق", 
                        "فروش موفق", "فروش ناموفق", "تعداد واحد", "مبلغ فروش"]
        
        for col, header in enumerate(daily_headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 2
        for date in sorted_dates:
            log_list = all_logs[date]
            if not isinstance(log_list, list):
                continue
            
            daily_visits = 0
            daily_successful = 0
            daily_failed = 0
            daily_sales_success = 0
            daily_sales_failed = 0
            daily_units = 0
            daily_amount = 0
            
            for log in log_list:
                if not isinstance(log, dict):
                    continue
                
                visit_status = log.get('visit_status', '')
                sales_status = log.get('sales_status', '')
                
                daily_visits += 1
                if visit_status == 'موفق':
                    daily_successful += 1
                    if sales_status == 'موفق':
                        daily_sales_success += 1
                        daily_units += safe_int(log.get('units_sold', 0))
                        daily_amount += safe_int(log.get('sales_amount', 0))
                    else:
                        daily_sales_failed += 1
                else:
                    daily_failed += 1
            
            ws3.cell(row=row, column=1, value=safe_str(date))
            ws3.cell(row=row, column=2, value=daily_visits)
            ws3.cell(row=row, column=3, value=daily_successful)
            ws3.cell(row=row, column=4, value=daily_failed)
            ws3.cell(row=row, column=5, value=daily_sales_success)
            ws3.cell(row=row, column=6, value=daily_sales_failed)
            ws3.cell(row=row, column=7, value=daily_units)
            ws3.cell(row=row, column=8, value=daily_amount)
            row += 1
        
        for col in range(1, len(daily_headers) + 1):
            ws3.column_dimensions[get_column_letter(col)].width = 15
        
        # ========== ذخیره فایل ==========
        reports_dir = os.path.join(data_path, 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        
        from datetime import datetime
        filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(reports_dir, filename)
        
        wb.save(filepath)
        
        return filepath
        
    except Exception as e:
        print(f"❌ خطا در خروجی Excel: {e}")
        return None